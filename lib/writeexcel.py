from config import setting
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import *
import configparser as cparser

cf = cparser.ConfigParser()
cf.read(setting.TEST_CONFIG, encoding='UTF-8')
result_index = int(cf.get('index', 'result'))+1
testers = int(cf.get('index', 'testers'))+1
name = cf.get('tester', 'name')


class WriteExcel:
    def __init__(self):
        self.filename = setting.TARGET_FILE
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def write_data(self, row, result):
        """
        写入测试结果
        :param row_n:数据所在行数
        :param value: 测试结果值
        :return: 无
        """
        row += 1
        font_GREEN = Font(name='宋体', color=DARKGREEN, bold=True)
        font_RED = Font(name='宋体', color=RED, bold=True)
        font1 = Font(name='宋体', color=DARKRED, bold=True)
        align = Alignment(horizontal='center', vertical='center')
        # 获数所在行数
        I_n = 'I' + str(row)
        J_n = 'J' + str(row)

        if result == True:
            self.ws.cell(row, result_index, 'PASS')
            self.ws[I_n].font = font_GREEN
        else:
            self.ws.cell(row, result_index, 'FAIL')
            self.ws[I_n].font = font_RED
        self.ws.cell(row, testers, name)
        self.ws[I_n].alignment = align
        self.ws[J_n].font = font1
        self.ws[J_n].alignment = align
        self.wb.save(self.filename)


if __name__ == '__main__':
    w = WriteExcel()
    w.write_data(5, 'PASS')
    w.write_data(6, True)